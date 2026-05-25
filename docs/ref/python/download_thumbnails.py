"""
縮圖批次下載腳本
從 卒業アルバム.xlsx 讀取所有影片 URL，下載縮圖並依分頁分類存放

使用方法：
  python download_thumbnails.py

輸出結構：
  thumbnails/
    メン限/
      abc123.jpg
      ...
    VALORANT/
      ...
"""

import os
import re
import time
import urllib.request
from openpyxl import load_workbook

# ── 設定 ──────────────────────────────────────────────────────────────────────
XLSX_PATH      = "卒業アルバム.xlsx"   # Excel ファイルのパス（同じフォルダに置く）
OUTPUT_DIR     = "thumbnails"          # 出力先フォルダ
SKIP_SHEETS    = ["🏠 トップ"]         # スキップするシート名
HEADER_ROW     = 1                     # ヘッダー行の番号（1始まり）
URL_COLUMN     = 3                     # URL が入っている列番号（C列 = 3）

# 縮圖解像度（上から順に試す）
RESOLUTIONS = [
    "maxresdefault",   # 最高画質 (1280x720)
    "hqdefault",       # 高画質   (480x360)
    "mqdefault",       # 中画質   (320x180)
]

# ── ユーティリティ ─────────────────────────────────────────────────────────────
def extract_video_id(url):
    """YouTube URL から動画 ID を抽出"""
    if not url:
        return None
    patterns = [
        r"youtu\.be/([A-Za-z0-9_\-]{11})",
        r"youtube\.com/watch\?v=([A-Za-z0-9_\-]{11})",
        r"youtube\.com/v/([A-Za-z0-9_\-]{11})",
    ]
    for pattern in patterns:
        m = re.search(pattern, str(url))
        if m:
            return m.group(1)
    return None

def safe_folder_name(name):
    """フォルダ名として使えない文字を除去"""
    return re.sub(r'[\\/*?:"<>|]', '_', name).strip()

def download_thumbnail(video_id, save_path):
    """縮圖をダウンロード（高解像度から順に試す）"""
    for res in RESOLUTIONS:
        url = f"https://img.youtube.com/vi/{video_id}/{res}.jpg"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                if resp.status == 200:
                    with open(save_path, "wb") as f:
                        f.write(resp.read())
                    return res   # 成功した解像度を返す
        except Exception:
            continue
    return None   # 全部失敗

# ── メイン ────────────────────────────────────────────────────────────────────
def main():
    print(f"📂 {XLSX_PATH} を読み込み中...")
    wb = load_workbook(XLSX_PATH, read_only=True)

    total_ok    = 0
    total_skip  = 0
    total_fail  = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        folder = os.path.join(OUTPUT_DIR, safe_folder_name(sheet_name))
        os.makedirs(folder, exist_ok=True)

        sheet_ok = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < HEADER_ROW:   # ヘッダー行をスキップ
                continue

            url = row[URL_COLUMN - 1] if len(row) >= URL_COLUMN else None
            video_id = extract_video_id(url)

            if not video_id:
                total_skip += 1
                continue
            
            # file name = YYMMDD Title Yt-ID
            if(0):
                save_path = os.path.join(folder, f"{video_id}.jpg")
            else:
                date  = str(row[0]).strip() if row[0] else "000000"
                title = str(row[1]).strip() if row[1] else "notitle"
                # ファイル名に使えない文字を除去
                title = re.sub(r'[\\/*?:"<>|\n\t]', '_', title)[:50]  # 長すぎる場合は50文字で切る
                save_path = os.path.join(folder, f"{date} {title} {video_id}.jpg")

            # 既にダウンロード済みならスキップ
            if os.path.exists(save_path):
                total_skip += 1
                continue

            res = download_thumbnail(video_id, save_path)
            if res:
                sheet_ok  += 1
                total_ok  += 1
                print(f"  ✅ {video_id}  [{res}]")
            else:
                total_fail += 1
                print(f"  ❌ {video_id}  (ダウンロード失敗)")

            time.sleep(0.3)   # サーバーへの負荷を抑える

        if sheet_ok > 0:
            print(f"🗂️  [{sheet_name}]  {sheet_ok} 枚完了\n")

    print("=" * 50)
    print(f"✅ 成功: {total_ok} 枚")
    print(f"⏭️  スキップ（既存 or URL なし）: {total_skip} 件")
    print(f"❌ 失敗: {total_fail} 件")
    print(f"📁 保存先: {os.path.abspath(OUTPUT_DIR)}/")

if __name__ == "__main__":
    main()
