import os
import re
from datetime import datetime
from googleapiclient.discovery import build
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 設定 ──────────────────────────────────────────────────────────────────────
API_KEY    = ""
CHANNEL_ID = "UChmCIfTk8UsT4igvKZ6Huqw"   # 結萌ひまり的 channel ID
OUTPUT     = "test.xlsx"

# ── 顏色 ──────────────────────────────────────────────────────────────────────
PINK_DARK  = "E87FAB"
PINK_LIGHT = "FFEEF5"
WHITE      = "FFFFFF"
TEXT_DARK  = "3D2B3D"
TEXT_MID   = "7A5C7A"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font():
    return Font(bold=True, size=10, color="FFFFFF", name="Yu Gothic")

def normal_font(color=TEXT_DARK):
    return Font(size=10, color=color, name="Yu Gothic")

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center")

# ── YouTube API ───────────────────────────────────────────────────────────────
def get_youtube():
    return build("youtube", "v3", developerKey=API_KEY)

def get_all_playlists(yt):
    """チャンネルの全プレイリストを取得"""
    playlists = []
    page_token = None
    while True:
        res = yt.playlists().list(
            part="snippet",
            channelId=CHANNEL_ID,
            maxResults=50,
            pageToken=page_token
        ).execute()
        for item in res.get("items", []):
            playlists.append({
                "id":    item["id"],
                "title": item["snippet"]["title"]
            })
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return playlists

def get_playlist_videos(yt, playlist_id):
    """プレイリスト内の全動画を取得"""
    videos = []
    page_token = None
    while True:
        res = yt.playlistItems().list(
            part="snippet,contentDetails",
            playlistId=playlist_id,
            maxResults=50,
            pageToken=page_token
        ).execute()
        for item in res.get("items", []):
            snippet = item["snippet"]
            video_id = snippet["resourceId"]["videoId"]

            # 公開日を YYMMDD に変換
            published = snippet.get("publishedAt", "")
            try:
                dt = datetime.strptime(published[:10], "%Y-%m-%d")
                date_str = dt.strftime("%y%m%d")
            except Exception:
                date_str = "??????"

            videos.append({
                "date":  date_str,
                "title": snippet.get("title", "（タイトル不明）"),
                "url":   f"https://youtu.be/{video_id}"
            })
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return videos

# ── Excel 出力 ────────────────────────────────────────────────────────────────
def write_excel(playlists_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # デフォルトシートを削除

    for pl in playlists_data:
        # シート名は31文字以内・特殊文字禁止
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', pl["title"])[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 列幅
        ws.column_dimensions["A"].width = 12   # 日付
        ws.column_dimensions["B"].width = 55   # タイトル
        ws.column_dimensions["C"].width = 40   # URL

        # ヘッダー行
        ws.row_dimensions[1].height = 22
        headers = ["配信日 (YYMMDD)", "タイトル", "URL"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font      = header_font()
            c.fill      = fill(PINK_DARK)
            c.alignment = center()

        # データ行
        for i, video in enumerate(pl["videos"]):
            r  = i + 2
            bg = PINK_LIGHT if i % 2 == 0 else WHITE
            ws.row_dimensions[r].height = 18

            # 日付
            c_date = ws.cell(row=r, column=1, value=video["date"])
            c_date.font      = normal_font(TEXT_MID)
            c_date.fill      = fill(bg)
            c_date.alignment = center()

            # タイトル
            c_title = ws.cell(row=r, column=2, value=video["title"])
            c_title.font      = normal_font()
            c_title.fill      = fill(bg)
            c_title.alignment = left()

            # URL（ハイパーリンク）
            c_url = ws.cell(row=r, column=3, value=video["url"])
            c_url.hyperlink  = video["url"]
            c_url.font       = Font(size=10, color="0563C1",
                                    underline="single", name="Yu Gothic")
            c_url.fill       = fill(bg)
            c_url.alignment  = left()

        ws.freeze_panes = "A2"
        ws.sheet_properties.tabColor = PINK_DARK

        print(f"  ✅ {pl['title']}  ({len(pl['videos'])} 本)")

    wb.save(OUTPUT)
    print(f"\n💾 保存完了：{OUTPUT}")

# ── メイン ────────────────────────────────────────────────────────────────────
def main():
    print("📡 YouTube API に接続中...")
    yt = get_youtube()

    print("📋 プレイリスト一覧を取得中...")
    playlists = get_all_playlists(yt)
    print(f"   {len(playlists)} 個のプレイリストを発見\n")

    results = []
    for pl in playlists:
        print(f"🎬 [{pl['title']}] の動画を取得中...")
        videos = get_playlist_videos(yt, pl["id"])
        results.append({"title": pl["title"], "videos": videos})

    print("\n📊 Excel ファイルを作成中...")
    write_excel(results)

if __name__ == "__main__":
    main()
