from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime

wb = Workbook()

# ── Color Palette (soft floral / himari theme) ──────────────────────────────
PINK_LIGHT   = "FFEEF5"   # background blocks
PINK_MID     = "F9C8DC"   # section headers
PINK_DARK    = "E87FAB"   # accent / borders
CREAM        = "FFFAF5"   # main bg
GREEN_SOFT   = "D4EDD4"   # stat blocks
GREEN_DARK   = "6BAD6B"   # stat accent
PURPLE_SOFT  = "EDE0F5"   # form link block
PURPLE_DARK  = "9B6FC8"   # purple accent
TEXT_DARK    = "3D2B3D"   # main text
TEXT_MID     = "7A5C7A"   # subtext
WHITE        = "FFFFFF"
YELLOW_SOFT  = "FFF8DC"   # version block

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=TEXT_DARK, italic=False, name="Yu Gothic"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color=PINK_DARK):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=PINK_DARK):
    return Border(bottom=Side(style="medium", color=color))

def set_cell(ws, row, col, value="", bold=False, size=11, color=TEXT_DARK,
             italic=False, bg=None, h_align="left", v_align="center",
             wrap=False, border=None, hyperlink=None, font_name="Yu Gothic"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color, italic=italic, name=font_name)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    if border:
        c.border = border
    if hyperlink:
        c.hyperlink = hyperlink
        c.font = Font(bold=bold, size=size, color="0563C1", italic=italic,
                      name=font_name, underline="single")
    return c

def merge_set(ws, r1, c1, r2, c2, value="", bold=False, size=11, color=TEXT_DARK,
              italic=False, bg=None, h_align="center", v_align="center",
              wrap=False, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = set_cell(ws, r1, c1, value, bold, size, color, italic, bg,
                 h_align, v_align, wrap, border)
    return c

# ════════════════════════════════════════════════════════════════════════════
# Sheet 1 : 導覽頁
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "🏠 トップ"

# Column widths
col_widths = {1: 2, 2: 18, 3: 22, 4: 22, 5: 22, 6: 18, 7: 2}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Row heights
for r in range(1, 60):
    ws.row_dimensions[r].height = 18
ws.row_dimensions[1].height = 8   # top margin
ws.row_dimensions[3].height = 42  # title
ws.row_dimensions[4].height = 22  # subtitle

# ── Full background fill ─────────────────────────────────────────────────────
for row in ws.iter_rows(min_row=1, max_row=58, min_col=1, max_col=7):
    for cell in row:
        cell.fill = fill(CREAM)

# ── Row 2: decorative top banner ─────────────────────────────────────────────
merge_set(ws, 2, 1, 2, 7, "🌸 ✿ 🌸 ✿ 🌸 ✿ 🌸 ✿ 🌸 ✿ 🌸 ✿ 🌸 ✿ 🌸",
          bold=False, size=10, color=PINK_DARK, bg=PINK_MID, h_align="center")

# ── Row 3: Main title ────────────────────────────────────────────────────────
merge_set(ws, 3, 1, 3, 7,
          "🏵️ 結萌ひまり 配信アーカイブ まとめ 🏵️",
          bold=True, size=20, color=TEXT_DARK, bg=PINK_LIGHT, h_align="center")

# ── Row 4: Subtitle ──────────────────────────────────────────────────────────
merge_set(ws, 4, 1, 4, 7,
          "Musubime Himari｜Re:AcT Gaming所属 Vtuber｜配信まとめ表",
          bold=False, size=10, color=TEXT_MID, bg=PINK_LIGHT, h_align="center",
          italic=True)

# ── Row 5: divider ───────────────────────────────────────────────────────────
merge_set(ws, 5, 1, 5, 7, "", bg=PINK_DARK)
ws.row_dimensions[5].height = 4

# ── Row 6: spacing ───────────────────────────────────────────────────────────
ws.row_dimensions[6].height = 10

# ── Rows 7-9: Channel Info ───────────────────────────────────────────────────
merge_set(ws, 7, 2, 7, 6, "📺  チャンネル基本情報",
          bold=True, size=12, color=WHITE, bg=PINK_DARK, h_align="left")

info_rows = [
    ("YouTubeチャンネル", "https://www.youtube.com/@raghimari", True),
    ("X (Twitter)",        "https://x.com/raghimari",            True),
    ("所属事務所",         "Re:AcT Gaming  /  https://www.v-react.com", False),
]
for i, (label, val, is_link) in enumerate(info_rows):
    r = 8 + i
    ws.row_dimensions[r].height = 20
    set_cell(ws, r, 2, f"  {label}", bold=True, size=10, color=TEXT_MID, bg=PINK_LIGHT)
    if is_link:
        set_cell(ws, r, 3, val, size=10, bg=PINK_LIGHT, hyperlink=val)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    else:
        set_cell(ws, r, 3, val, size=10, bg=PINK_LIGHT)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)

ws.row_dimensions[11].height = 10

# ── Row 12: Stats section header ─────────────────────────────────────────────
merge_set(ws, 12, 2, 12, 6, "📊  統計データ",
          bold=True, size=12, color=WHITE, bg=GREEN_DARK, h_align="left")

# ── Rows 13-14: Stat boxes ───────────────────────────────────────────────────
ws.row_dimensions[13].height = 30
ws.row_dimensions[14].height = 20

stat_labels = ["📁 プレイリスト数", "🎬 動画総数", "📅 最終更新日"]
stat_values = ["※ 手動填寫", "※ 手動填寫", "※ 手動填寫"]
stat_cols   = [2, 3, 4]   # each takes ~1.5 cols; we'll merge pairs

stat_col_spans = [(2,3), (3,4), (4,5), (5,6)]  # 4 blocks across cols 2-6
for j, (label, val) in enumerate(zip(stat_labels, stat_values)):
    c1, c2 = [(2,3),(3,4),(5,6)][j]  # manual spans
# Simpler: 3 equal stat boxes across B-F
for j in range(3):
    c_start = 2 + j
    set_cell(ws, 13, c_start, stat_labels[j], bold=True, size=9,
             color=GREEN_DARK, bg=GREEN_SOFT, h_align="center", v_align="bottom")
    set_cell(ws, 14, c_start, stat_values[j], bold=False, size=9,
             color=TEXT_MID, bg=GREEN_SOFT, h_align="center", v_align="top",
             italic=True)

# fill col 5 with same bg
set_cell(ws, 13, 5, "", bg=GREEN_SOFT)
set_cell(ws, 14, 5, "", bg=GREEN_SOFT)
set_cell(ws, 13, 6, "", bg=GREEN_SOFT)
set_cell(ws, 14, 6, "", bg=GREEN_SOFT)

ws.row_dimensions[15].height = 10

# ── Row 16: Playlist nav header ──────────────────────────────────────────────
merge_set(ws, 16, 2, 16, 6, "🗂️  プレイリスト一覧（クリックで移動）",
          bold=True, size=12, color=WHITE, bg=PINK_DARK, h_align="left")

# ── Rows 17-21: Playlist links (5 placeholder rows) ─────────────────────────
playlist_examples = [
    ("🎮 ストリートファイター6",    "SF6"),
    ("👾 バイオハザード",           "Biohazard"),
    ("🌿 マインクラフト",           "Minecraft"),
    ("✨ 雑談・歌配信",             "Zatsudan"),
    ("📦 その他",                   "Other"),
]
for i, (name, sheet_hint) in enumerate(playlist_examples):
    r = 17 + i
    ws.row_dimensions[r].height = 22
    bg = PINK_LIGHT if i % 2 == 0 else WHITE
    merge_set(ws, r, 2, r, 3, f"  {name}", bold=False, size=11,
              color=TEXT_DARK, bg=bg, h_align="left")
    # internal hyperlink placeholder
    note_cell = set_cell(ws, r, 4, f"→ #{sheet_hint} シートへ",
                         size=10, color=TEXT_MID, bg=bg, h_align="center",
                         italic=True)
    merge_set(ws, r, 4, r, 6, f"→ ※ シート名を設定後にリンクを追加",
              bold=False, size=9, color=TEXT_MID, bg=bg, h_align="right",
              italic=True)

ws.row_dimensions[22].height = 10

# ── Row 23: Form link block ───────────────────────────────────────────────────
merge_set(ws, 23, 2, 23, 6, "📝  情報提供フォーム",
          bold=True, size=12, color=WHITE, bg=PURPLE_DARK, h_align="left")

ws.row_dimensions[24].height = 22
ws.row_dimensions[25].height = 20
merge_set(ws, 24, 2, 24, 6,
          "アーカイブ情報の補足・修正にご協力ください！",
          bold=False, size=10, color=TEXT_DARK, bg=PURPLE_SOFT, h_align="center")
set_cell(ws, 25, 2, "  フォームURL：", bold=True, size=10,
         color=PURPLE_DARK, bg=PURPLE_SOFT)
form_url = "https://forms.gle/xxxxxxxxxxxxxxxx"   # placeholder
set_cell(ws, 25, 3, form_url, size=10, bg=PURPLE_SOFT,
         hyperlink=form_url)
ws.merge_cells(start_row=25, start_column=3, end_row=25, end_column=6)

ws.row_dimensions[26].height = 10

# ── Row 27: Version log header ────────────────────────────────────────────────
merge_set(ws, 27, 2, 27, 6, "🗒️  更新履歴 / バージョン記録",
          bold=True, size=12, color=WHITE, bg="A08070", h_align="left")

ws.row_dimensions[28].height = 18
set_cell(ws, 28, 2, "  バージョン", bold=True, size=9, color=TEXT_MID, bg=YELLOW_SOFT)
set_cell(ws, 28, 3, "更新日",       bold=True, size=9, color=TEXT_MID, bg=YELLOW_SOFT, h_align="center")
ws.merge_cells(start_row=28, start_column=4, end_row=28, end_column=6)
set_cell(ws, 28, 4, "変更内容",     bold=True, size=9, color=TEXT_MID, bg=YELLOW_SOFT, h_align="center")

ver_rows = [
    ("v0.1", str(datetime.date.today()), "初版作成・導覧ページ設置"),
    ("",     "",                          ""),
    ("",     "",                          ""),
]
for i, (ver, date, note) in enumerate(ver_rows):
    r = 29 + i
    ws.row_dimensions[r].height = 18
    bg = YELLOW_SOFT if i % 2 == 0 else WHITE
    set_cell(ws, r, 2, f"  {ver}", size=10, color=TEXT_DARK, bg=bg)
    set_cell(ws, r, 3, date,       size=10, color=TEXT_DARK, bg=bg, h_align="center")
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    set_cell(ws, r, 4, note,       size=10, color=TEXT_DARK, bg=bg)

# ── Bottom banner ────────────────────────────────────────────────────────────
ws.row_dimensions[33].height = 4
merge_set(ws, 33, 1, 33, 7, "", bg=PINK_DARK)
merge_set(ws, 34, 1, 34, 7,
          "🌸 ✿ 作成者: ※ あなたのお名前 ✿  |  ✿ ご自由にご利用ください ✿ 🌸",
          bold=False, size=9, color=TEXT_MID, bg=PINK_LIGHT, h_align="center",
          italic=True)

# ── Freeze top rows ──────────────────────────────────────────────────────────
ws.freeze_panes = "B6"

# ── Tab color ────────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = "E87FAB"

# ════════════════════════════════════════════════════════════════════════════
# Sample playlist sheet template (×1 as example)
# ════════════════════════════════════════════════════════════════════════════
for sheet_name, tab_color in [
    ("🎮 SF6",        "E87FAB"),
    ("👾 Biohazard",  "A08070"),
    ("🌿 Minecraft",  "6BAD6B"),
    ("✨ 雑談・歌",   "9B6FC8"),
    ("📦 その他",     "C8A06B"),
]:
    ws2 = wb.create_sheet(sheet_name)
    ws2.sheet_properties.tabColor = tab_color

    # column widths
    ws2.column_dimensions["A"].width = 12   # date
    ws2.column_dimensions["B"].width = 52   # title
    ws2.column_dimensions["C"].width = 40   # URL

    # header banner
    ws2.row_dimensions[1].height = 8
    ws2.row_dimensions[2].height = 30
    ws2.row_dimensions[3].height = 4
    ws2.row_dimensions[4].height = 20

    ws2.merge_cells("A2:C2")
    set_cell(ws2, 2, 1, f"🏵️  {sheet_name.replace('🎮 ','').replace('👾 ','').replace('🌿 ','').replace('✨ ','').replace('📦 ','')}  アーカイブ",
             bold=True, size=15, color=TEXT_DARK, bg=PINK_LIGHT, h_align="center")

    ws2.merge_cells("A3:C3")
    set_cell(ws2, 3, 1, "", bg=PINK_DARK)

    headers = ["配信日 (YYMMDD)", "タイトル", "URL"]
    for j, h in enumerate(headers):
        c = ws2.cell(row=4, column=j+1, value=h)
        c.font = Font(bold=True, size=10, color=WHITE, name="Yu Gothic")
        c.fill = fill(PINK_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Sample rows
    for r in range(5, 15):
        ws2.row_dimensions[r].height = 18
        bg = PINK_LIGHT if r % 2 == 1 else WHITE
        for col in range(1, 4):
            ws2.cell(row=r, column=col).fill = fill(bg)

    ws2.freeze_panes = "A5"

    # Back to nav link
    ws2.row_dimensions[16].height = 20
    set_cell(ws2, 16, 1, "← 🏠 トップへ戻る", size=9, color=PURPLE_DARK,
             italic=True, hyperlink="#'🏠 トップ'!A1")

wb.save("/home/claude/himari_archive.xlsx")
print("saved!")
