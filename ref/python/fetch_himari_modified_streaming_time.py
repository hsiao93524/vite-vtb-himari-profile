"""
結萌ひまり 配信アーカイブ
修正版：直播實際開始時間取得腳本

主要改善：
  - 不再使用 playlistItems 的 publishedAt（那是「加入清單的日期」）
  - 改用 videos().list() 取得 liveStreamingDetails.actualStartTime（直播實際開始時間）
  - 普通影片 / ショート則 fallback 到 snippet.publishedAt（影片公開日期）

使用方法：
  pip install google-api-python-client openpyxl
  python fetch_himari_modified_streaming_time.py
"""

import os
import re
import time
from datetime import datetime, timezone

from googleapiclient.discovery import build
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 設定 ──────────────────────────────────────────────────────────────────────
API_KEY    = "AIzaSyABY8X9SR2aADEmthdPWwNYNKW9AFDnFDc"
CHANNEL_ID = "UChmCIfTk8UsT4igvKZ6Huqw"   # 結萌ひまり
OUTPUT     = "卒業アルバム_修正版.xlsx"

# ── 顏色 ──────────────────────────────────────────────────────────────────────
PINK_DARK  = "E87FAB"
PINK_LIGHT = "FFEEF5"
WHITE      = "FFFFFF"
TEXT_DARK  = "3D2B3D"
TEXT_MID   = "7A5C7A"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font():
    return Font(bold=True, size=10, color="FFFFFF", name="Yu Gothic")

def normal_font(color=TEXT_DARK):
    return Font(size=10, color=color, name="Yu Gothic")

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center")

# ── YouTube API ───────────────────────────────────────────────────────────────
def get_youtube():
    return build("youtube", "v3", developerKey=API_KEY)


def get_all_playlists(yt):
    """チャンネルの全プレイリストを取得"""
    playlists = []
    page_token = None
    while True:
        res = yt.playlists().list(
            part="snippet",
            channelId=CHANNEL_ID,
            maxResults=50,
            pageToken=page_token
        ).execute()
        for item in res.get("items", []):
            playlists.append({
                "id":    item["id"],
                "title": item["snippet"]["title"]
            })
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return playlists


def get_playlist_video_ids(yt, playlist_id):
    """
    プレイリスト内の全動画の (videoId, title) を取得
    日付はここでは取得しない（videos().list() で後から正確な日付を取る）
    """
    items = []
    page_token = None
    while True:
        res = yt.playlistItems().list(
            part="snippet,contentDetails",
            playlistId=playlist_id,
            maxResults=50,
            pageToken=page_token
        ).execute()
        for item in res.get("items", []):
            snippet  = item["snippet"]
            video_id = snippet["resourceId"]["videoId"]
            title    = snippet.get("title", "（タイトル不明）")
            url      = f"https://youtu.be/{video_id}"
            items.append({
                "videoId": video_id,
                "title":   title,
                "url":     url,
                "date":    None,   # 後で埋める
            })
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return items


def fetch_video_details(yt, video_ids: list[str]) -> dict[str, dict]:
    """
    videos().list() で正確な日付を取得（50件ずつバッチ処理）

    優先順位:
      1. liveStreamingDetails.actualStartTime  （直播實際開始時間 ← 最正確）
      2. liveStreamingDetails.scheduledStartTime（預定開始時間，直播但沒有實際時間時）
      3. snippet.publishedAt                   （影片公開時間，ショート/通常動画はこれ）
    """
    result = {}
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i : i + 50]
        res = yt.videos().list(
            part="snippet,liveStreamingDetails",
            id=",".join(batch),
            maxResults=50
        ).execute()

        for item in res.get("items", []):
            vid     = item["id"]
            snippet = item.get("snippet", {})
            live    = item.get("liveStreamingDetails", {})

            # 優先順位で日付を決定
            raw_dt = (
                live.get("actualStartTime")       # 直播實際開始 ← 最優先
                or live.get("scheduledStartTime") # 預定開始時間
                or snippet.get("publishedAt")     # 影片公開日（フォールバック）
            )

            if raw_dt:
                try:
                    # ISO8601 → YYMMDD
                    dt = datetime.fromisoformat(raw_dt.replace("Z", "+00:00"))
                    date_str = dt.strftime("%y%m%d")
                    date_source = (
                        "actualStart"   if live.get("actualStartTime")
                        else "scheduled" if live.get("scheduledStartTime")
                        else "published"
                    )
                except Exception:
                    date_str    = "??????"
                    date_source = "error"
            else:
                date_str    = "??????"
                date_source = "none"

            result[vid] = {
                "date":        date_str,
                "date_source": date_source,  # デバッグ用
            }

        print(f"  動画詳細取得: {min(i + 50, len(video_ids))}/{len(video_ids)} 件")
        time.sleep(0.1)   # API レート制限対策

    return result


# ── Excel 出力 ────────────────────────────────────────────────────────────────
def write_excel(playlists_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for pl in playlists_data:
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', pl["title"])[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 列幅（date_source デバッグ列を追加）
        ws.column_dimensions["A"].width = 12   # 日付
        ws.column_dimensions["B"].width = 55   # タイトル
        ws.column_dimensions["C"].width = 40   # URL
        ws.column_dimensions["D"].width = 14   # date_source（デバッグ用）

        # ヘッダー行
        ws.row_dimensions[1].height = 22
        headers = ["配信日 (YYMMDD)", "タイトル", "URL", "日付ソース"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font      = header_font()
            c.fill      = fill(PINK_DARK)
            c.alignment = center()

        # データ行
        for i, video in enumerate(pl["videos"]):
            r  = i + 2
            bg = PINK_LIGHT if i % 2 == 0 else WHITE
            ws.row_dimensions[r].height = 18

            # 日付
            c_date = ws.cell(row=r, column=1, value=video["date"])
            c_date.font      = normal_font(TEXT_MID)
            c_date.fill      = fill(bg)
            c_date.alignment = center()

            # タイトル
            c_title = ws.cell(row=r, column=2, value=video["title"])
            c_title.font      = normal_font()
            c_title.fill      = fill(bg)
            c_title.alignment = left()

            # URL（ハイパーリンク）
            c_url = ws.cell(row=r, column=3, value=video["url"])
            c_url.hyperlink  = video["url"]
            c_url.font       = Font(size=10, color="0563C1",
                                    underline="single", name="Yu Gothic")
            c_url.fill       = fill(bg)
            c_url.alignment  = left()

            # date_source（デバッグ用：どこから日付を取ったか）
            source_color = {
                "actualStart": "2E7D32",   # 濃い緑 = 最も正確
                "scheduled":   "F57F17",   # オレンジ = 予定日
                "published":   "1565C0",   # 青 = 公開日（フォールバック）
                "error":       "B71C1C",   # 赤 = エラー
                "none":        "9E9E9E",   # グレー = 取得できず
            }.get(video.get("date_source", "none"), "9E9E9E")

            c_src = ws.cell(row=r, column=4,
                            value=video.get("date_source", ""))
            c_src.font      = Font(size=9, color=source_color, name="Yu Gothic")
            c_src.fill      = fill(bg)
            c_src.alignment = center()

        ws.freeze_panes = "A2"
        ws.sheet_properties.tabColor = PINK_DARK

        print(f"  ✅ {pl['title']}  ({len(pl['videos'])} 本)")

    wb.save(OUTPUT)
    print(f"\n💾 保存完了：{OUTPUT}")


# ── メイン ────────────────────────────────────────────────────────────────────
def main():
    print("📡 YouTube API に接続中...")
    yt = get_youtube()

    print("📋 プレイリスト一覧を取得中...")
    playlists = get_all_playlists(yt)
    print(f"   {len(playlists)} 個のプレイリストを発見\n")

    # Step 1: 全プレイリストから videoId を収集
    results = []
    all_video_ids = []
    for pl in playlists:
        print(f"🎬 [{pl['title']}] の動画IDを取得中...")
        videos = get_playlist_video_ids(yt, pl["id"])
        results.append({"title": pl["title"], "videos": videos})
        all_video_ids.extend(v["videoId"] for v in videos if v["videoId"])

    # 重複除去（同じ動画が複数プレイリストにある場合）
    unique_ids = list(dict.fromkeys(all_video_ids))
    print(f"\n📡 動画詳細（正確な配信日）を取得中... 合計 {len(unique_ids)} 件")

    # Step 2: videos().list() で正確な日付を一括取得
    detail_map = fetch_video_details(yt, unique_ids)

    # Step 3: 各動画に日付を反映
    for pl_data in results:
        for video in pl_data["videos"]:
            vid = video["videoId"]
            if vid and vid in detail_map:
                video["date"]        = detail_map[vid]["date"]
                video["date_source"] = detail_map[vid]["date_source"]
            else:
                video["date"]        = "??????"
                video["date_source"] = "none"

    # Step 4: Excel 出力
    print("\n📊 Excel ファイルを作成中...")
    write_excel(results)

    # Step 5: サマリー表示
    source_counts = {}
    for pl_data in results:
        for v in pl_data["videos"]:
            src = v.get("date_source", "none")
            source_counts[src] = source_counts.get(src, 0) + 1

    print("\n📈 日付ソース サマリー:")
    labels = {
        "actualStart": "🟢 actualStartTime（直播實際開始）",
        "scheduled":   "🟡 scheduledStartTime（預定開始）",
        "published":   "🔵 publishedAt（公開日フォールバック）",
        "error":       "🔴 エラー",
        "none":        "⚪ 取得できず",
    }
    for src, label in labels.items():
        if src in source_counts:
            print(f"  {label}: {source_counts[src]} 件")


if __name__ == "__main__":
    main()
